"""Executor — single-threaded query plan execution with cursor management.

The Executor is the heart of the query engine. It takes a compiled QueryPlan
and executes it against a live ReplayController, returning a list of result
rows (dicts). Key responsibilities:

1. Materialize rows from the Catalog based on the plan's source spec.
2. Apply transformation steps (filter, sort, take, take_percent, with_counter)
   in the order specified by the plan — no implicit reordering.
3. Apply projection: for metadata fields, interpolate templates; for artifact
   fields, group by eventId, call SetFrameEvent once per group, then render
   each artifact to a file and return the file path.

The Executor is single-threaded by design (ReplayController requires thread
affinity). When busy, new execute() calls raise ExecutorBusy immediately.
"""

from __future__ import annotations

import os
import struct
import math
from typing import Any

import renderdoc as rd

from .catalog import Catalog
from .plan import QueryPlan, Step, Projection


class ExecutorBusy(Exception):
    """Raised when the executor is already running a query."""


class Executor:
    """Single-threaded query executor."""

    def __init__(self, controller: rd.ReplayController):
        self._catalog = Catalog(controller)
        self._busy = False
        self._current_event = None

    @property
    def catalog(self) -> Catalog:
        return self._catalog

    @property
    def is_busy(self) -> bool:
        return self._busy

    def execute(self, plan: QueryPlan, output_dir: str = "") -> list[dict]:
        """Execute a query plan and return result rows."""
        if self._busy:
            raise ExecutorBusy("executor is busy")

        out = output_dir or plan.output_dir
        self._busy = True
        try:
            rows = self._materialize_rows(plan)
            rows = self._apply_steps(plan.steps, rows)
            rows = self._apply_projection(plan.projection, rows, out)
            return rows
        finally:
            self._busy = False

    # ------------------------------------------------------------------
    # Row materialization
    # ------------------------------------------------------------------

    def _materialize_rows(self, plan: QueryPlan) -> list[dict]:
        """Load rows from the catalog based on the plan's source spec."""
        kind = plan.source.kind
        params = plan.source.params

        if kind == "actions":
            flags = params.get("flags")
            return [dict(r) for r in self._catalog.get_actions(flags)]
        elif kind == "events":
            return [dict(r) for r in self._catalog.events]
        elif kind == "resources":
            return [dict(r) for r in self._catalog.resources]
        else:
            raise ValueError(f"unknown source kind: {kind}")

    # ------------------------------------------------------------------
    # Step application (faithful ordering, no reordering)
    # ------------------------------------------------------------------

    def _apply_steps(self, steps: list[Step], rows: list[dict]) -> list[dict]:
        for step in steps:
            rows = self._apply_step(step, rows)
        return rows

    def _apply_step(self, step: Step, rows: list[dict]) -> list[dict]:
        op = step.op
        p = step.params

        if op == "with_counter":
            counter = p["counter"]
            counter_map = self._catalog.fetch_gpu_counters(counter)
            field_name = self._counter_field_name(counter)
            for row in rows:
                row[field_name] = counter_map.get(row["event_id"])
            return rows

        elif op == "filter":
            expr = p["expr"]
            return [r for r in rows if self._eval_predicate(expr, r)]

        elif op == "sort":
            field = p["field"]
            desc = p.get("desc", False)
            return sorted(rows, key=lambda r: self._sort_key(r, field), reverse=desc)

        elif op == "take":
            n = p["n"]
            return rows[:n]

        elif op == "take_percent":
            pct = p["pct"]
            count = max(1, math.ceil(len(rows) * pct / 100.0))
            return rows[:count]

        elif op == "group_by":
            # Group-by is a no-op for now; it marks the boundary for future
            # aggregate functions. Rows pass through unchanged.
            return rows

        else:
            raise ValueError(f"unknown step op: {op}")

    def _eval_predicate(self, expr: str, row: dict) -> bool:
        """Evaluate a filter predicate against a row.

        The expression can reference any row field by name (e.g. ``duration_gpu > 1000``).
        Uses a restricted eval with only the row's fields in scope.
        """
        safe_globals = {"__builtins__": {}}
        safe_locals = dict(row)
        try:
            return bool(eval(expr, safe_globals, safe_locals))
        except Exception:
            return False

    def _sort_key(self, row: dict, field: str):
        val = row.get(field)
        if val is None:
            return float("inf")  # None sorts last in ascending
        return val

    def _counter_field_name(self, counter: int) -> str:
        """Map a GPUCounter enum value to a field name."""
        # Generic counters have known names; fall back to f"counter_{counter}"
        counter_names = {
            int(rd.GPUCounter.EventGPUDuration): "duration_gpu",
            int(rd.GPUCounter.InputVerticesRead): "input_vertices_read",
            int(rd.GPUCounter.IAPrimitives): "ia_primitives",
            int(rd.GPUCounter.GSPrimitives): "gs_primitives",
            int(rd.GPUCounter.RasterizerInvocations): "rasterizer_invocations",
            int(rd.GPUCounter.RasterizedPrimitives): "rasterized_primitives",
            int(rd.GPUCounter.SamplesPassed): "samples_passed",
            int(rd.GPUCounter.VSInvocations): "vs_invocations",
            int(rd.GPUCounter.HSInvocations): "hs_invocations",
            int(rd.GPUCounter.DSInvocations): "ds_invocations",
            int(rd.GPUCounter.GSInvocations): "gs_invocations",
            int(rd.GPUCounter.PSInvocations): "ps_invocations",
            int(rd.GPUCounter.CSInvocations): "cs_invocations",
            int(rd.GPUCounter.ASInvocations): "as_invocations",
            int(rd.GPUCounter.MSInvocations): "ms_invocations",
        }
        return counter_names.get(counter, f"counter_{counter}")

    # ------------------------------------------------------------------
    # Projection (metadata interpolation + artifact materialization)
    # ------------------------------------------------------------------

    def _apply_projection(
        self, projections: list[Projection], rows: list[dict], output_dir: str
    ) -> list[dict]:
        if not projections:
            return rows

        # Separate metadata and artifact projections
        meta_projs = [p for p in projections if not p.is_artifact]
        art_projs = [p for p in projections if p.is_artifact]

        # Interpolate metadata fields
        for row in rows:
            for proj in meta_projs:
                row[proj.name] = self._interpolate(proj.expr, row)

        # Materialize artifacts (grouped by eventId for cursor efficiency)
        if art_projs and output_dir:
            os.makedirs(output_dir, exist_ok=True)
            self._materialize_artifacts(art_projs, rows, output_dir)

        # If no output_dir was set but artifacts were requested, skip silently
        # (artifacts are optional; metadata fields are still returned)

        # Trim to only projected fields
        all_names = {p.name for p in projections}
        if all_names:
            result = []
            for row in rows:
                result.append({k: row.get(k) for k in all_names})
            return result

        return rows

    def _interpolate(self, expr: str, row: dict) -> Any:
        """Interpolate a template like ``"{event_id}"`` or ``"{name}"`` from a row."""
        if expr.startswith("{") and expr.endswith("}"):
            field = expr[1:-1]
            return row.get(field)
        return expr

    def _materialize_artifacts(
        self, art_projs: list[Projection], rows: list[dict], output_dir: str
    ) -> None:
        """Group artifacts by eventId, set cursor once, then render all artifacts."""
        # Build a list of (event_id, projection, row) tuples
        tasks = []
        for row in rows:
            eid = row.get("event_id")
            if eid is None:
                continue
            for proj in art_projs:
                tasks.append((eid, proj, row))

        # Group by eventId to minimize SetFrameEvent calls
        current_eid = None
        for eid, proj, row in tasks:
            if eid != current_eid:
                self._catalog.controller.SetFrameEvent(eid, True)
                current_eid = eid

            path = self._render_artifact(proj, eid, row, output_dir)
            row[proj.name] = path

    def _render_artifact(
        self, proj: Projection, event_id: int, row: dict, output_dir: str
    ) -> str:
        """Render a single artifact to file and return the path."""
        kind = proj.expr
        params = proj.artifact_params
        ctrl = self._catalog.controller

        filename = f"event{event_id}_{proj.name}"
        out_path = os.path.join(output_dir, filename)

        if kind == "screenshot":
            return self._render_screenshot(params, event_id, out_path)
        elif kind == "mesh":
            return self._render_mesh(params, event_id, out_path)
        elif kind == "mesh_screenshot":
            return self._render_mesh_screenshot(params, event_id, out_path)
        elif kind == "texture_data":
            return self._render_texture_data(params, event_id, out_path)
        elif kind == "shader_disasm":
            return self._render_shader_disasm(params, event_id, out_path)
        elif kind == "buffer_data":
            return self._render_buffer_data(params, event_id, out_path)
        elif kind == "log":
            return self._render_log(params, event_id, out_path, row)
        elif kind == "excel":
            return self._render_excel(params, event_id, out_path, row)
        else:
            raise ValueError(f"unknown artifact kind: {kind}")

    # ------------------------------------------------------------------
    # Artifact renderers
    # ------------------------------------------------------------------

    def _render_screenshot(self, params: dict, event_id: int, out_path: str) -> str:
        """Capture the current output target (or specified texture) as an image."""
        ctrl = self._catalog.controller
        filetype_map = {
            "png": rd.FileType.PNG,
            "jpg": rd.FileType.JPG,
            "tga": rd.FileType.TGA,
            "bmp": rd.FileType.BMP,
        }
        ftype = filetype_map.get(params.get("filetype", "png"), rd.FileType.PNG)

        # Determine the texture to capture
        tex_id = params.get("texture_id")
        if tex_id:
            rid = _str_to_resource_id(tex_id)
        else:
            # Use the first color output target
            pipe = ctrl.GetPipelineState()
            targets = pipe.GetOutputTargets()
            if not targets or targets[0].resource == rd.ResourceId.Null():
                return ""
            rid = targets[0].resource

        save = rd.TextureSave()
        save.resourceId = rid
        save.destType = ftype
        save.mip = params.get("mip", 0)
        save.slice.sliceIndex = params.get("slice", 0)
        save.sample.sampleIndex = params.get("sample", 0)

        ext = {"png": ".png", "jpg": ".jpg", "tga": ".tga", "bmp": ".bmp"}[params.get("filetype", "png")]
        full_path = out_path + ext

        result = ctrl.SaveTexture(save, full_path)
        if result.OK():
            return full_path
        return ""

    def _render_mesh(self, params: dict, event_id: int, out_path: str) -> str:
        """Extract post-transform vertex data and save as OBJ."""
        ctrl = self._catalog.controller
        stage_map = {
            "PostVS": rd.MeshDataStage.VSOut,
            "PostGS": rd.MeshDataStage.GSOut,
            "PostMesh": rd.MeshDataStage.MeshOut,
            "TaskOut": rd.MeshDataStage.TaskOut,
        }
        stage = stage_map.get(params.get("stage", "PostVS"), rd.MeshDataStage.VSOut)
        instance = params.get("instance", 0)
        view = params.get("view", 0)

        mesh = ctrl.GetPostVSData(instance, view, stage)
        if mesh.numIndices == 0:
            return ""

        fmt = params.get("format", "obj")
        full_path = out_path + f".{fmt}"

        # Fetch vertex data from the post-VS buffer
        vert_data = ctrl.GetBufferData(mesh.vertexResourceId, mesh.vertexByteOffset, 0)
        if not vert_data:
            return ""

        if fmt == "obj":
            self._write_obj(full_path, mesh, vert_data)
        else:
            # For non-OBJ, write raw bytes as .bin
            with open(full_path, "wb") as f:
                f.write(vert_data)

        return full_path

    def _write_obj(self, path: str, mesh: rd.MeshFormat, vert_data: bytes) -> None:
        """Write vertex positions as a minimal OBJ file."""
        stride = mesh.vertexByteStride
        if stride <= 0:
            return

        # Position is always the first attribute in post-VS data
        # with 4 components (x, y, z, w) at 4-byte float
        pos_offset = 0
        pos_size = 12  # x, y, z = 3 floats = 12 bytes

        num_verts = len(vert_data) // stride

        with open(path, "w") as f:
            f.write("# RenderQuery mesh export\n")
            f.write(f"# {num_verts} vertices, stride={stride}\n")
            for i in range(num_verts):
                base = i * stride + pos_offset
                if base + pos_size > len(vert_data):
                    break
                x, y, z = struct.unpack_from("<fff", vert_data, base)
                f.write(f"v {x:.6f} {y:.6f} {z:.6f}\n")
            # Write a simple face referencing all vertices
            if num_verts >= 3:
                for i in range(0, num_verts - 2, 3):
                    f.write(f"f {i+1} {i+2} {i+3}\n")

    def _render_texture_data(self, params: dict, event_id: int, out_path: str) -> str:
        """Extract raw texture data and save to file."""
        ctrl = self._catalog.controller
        rid = _str_to_resource_id(params["resource_id"])
        sub = rd.Subresource(
            mip=params.get("mip", 0),
            slice=params.get("slice", 0),
            sample=params.get("sample", 0),
        )
        data = ctrl.GetTextureData(rid, sub)
        if not data:
            return ""

        filetype = params.get("filetype", "dds")
        if filetype == "dds":
            full_path = out_path + ".dds"
            with open(full_path, "wb") as f:
                f.write(data)
        else:
            # Use SaveTexture for image formats
            ftype_map = {
                "png": rd.FileType.PNG,
                "tga": rd.FileType.TGA,
                "bmp": rd.FileType.BMP,
            }
            save = rd.TextureSave()
            save.resourceId = rid
            save.destType = ftype_map.get(filetype, rd.FileType.PNG)
            save.mip = sub.mip
            save.slice.sliceIndex = sub.slice
            save.sample.sampleIndex = sub.sample
            full_path = out_path + f".{filetype}"
            result = ctrl.SaveTexture(save, full_path)
            if not result.OK():
                return ""

        return full_path

    def _render_shader_disasm(self, params: dict, event_id: int, out_path: str) -> str:
        """Extract shader disassembly and save as text."""
        ctrl = self._catalog.controller
        stage_map = {
            "Vertex": rd.ShaderStage.Vertex,
            "Pixel": rd.ShaderStage.Pixel,
            "Geometry": rd.ShaderStage.Geometry,
            "Hull": rd.ShaderStage.Hull,
            "Domain": rd.ShaderStage.Domain,
            "Compute": rd.ShaderStage.Compute,
            "Mesh": rd.ShaderStage.Mesh,
            "Amplification": rd.ShaderStage.Amplification,
        }
        stage = stage_map.get(params.get("stage", "Vertex"), rd.ShaderStage.Vertex)
        target = params.get("target", "")

        pipe = ctrl.GetPipelineState()
        shader_obj = pipe.GetShaderObject(stage)
        if shader_obj == rd.ResourceId.Null():
            return ""

        # Get shader reflection
        shader_refl = ctrl.GetShader(rd.ResourceId.Null(), shader_obj, rd.ShaderEntryPoint())
        if shader_refl is None:
            return ""

        disasm = ctrl.DisassembleShader(rd.ResourceId.Null(), shader_refl, target)
        if not disasm:
            return ""

        full_path = out_path + ".txt"
        with open(full_path, "w") as f:
            f.write(disasm)
        return full_path

    def _render_buffer_data(self, params: dict, event_id: int, out_path: str) -> str:
        """Extract raw buffer data and save as binary."""
        ctrl = self._catalog.controller
        rid = _str_to_resource_id(params["resource_id"])
        offset = params.get("offset", 0)
        length = params.get("length", 0)

        data = ctrl.GetBufferData(rid, offset, length)
        if not data:
            return ""

        full_path = out_path + ".bin"
        with open(full_path, "wb") as f:
            f.write(data)
        return full_path

    # ------------------------------------------------------------------
    # Mesh screenshot — render mesh via ReplayOutput with MeshDisplay
    # ------------------------------------------------------------------

    def _render_mesh_screenshot(self, params: dict, event_id: int, out_path: str) -> str:
        """Render the drawcall's mesh with wireframe highlight via headless ReplayOutput."""
        ctrl = self._catalog.controller
        width = params.get("width", 512)
        height = params.get("height", 512)
        stage_map = {
            "PostVS": rd.MeshDataStage.VSOut,
            "PostGS": rd.MeshDataStage.GSOut,
            "PostMesh": rd.MeshDataStage.MeshOut,
            "TaskOut": rd.MeshDataStage.TaskOut,
        }
        stage = stage_map.get(params.get("stage", "PostVS"), rd.MeshDataStage.VSOut)
        instance = params.get("instance", 0)
        view = params.get("view", 0)
        wireframe = params.get("wireframe", True)

        # Get post-VS mesh data (must be called after SetFrameEvent)
        mesh_fmt = ctrl.GetPostVSData(instance, view, stage)
        if mesh_fmt.numIndices == 0:
            return ""

        # Create headless mesh output
        window = rd.CreateHeadlessWindowingData(width, height)
        output = ctrl.CreateOutput(window, rd.ReplayOutputType.Mesh)

        # Configure mesh display (following rdtest Mesh_Zoo pattern)
        cfg = rd.MeshDisplay()
        cfg.type = stage
        cfg.cam = rd.InitCamera(rd.CameraType.FPSLook)
        cfg.position = mesh_fmt
        cfg.position.nearPlane = 1.0
        cfg.position.farPlane = 100.0
        cfg.ortho = True
        cfg.aspect = float(width) / float(height)
        cfg.wireframeDraw = wireframe
        cfg.position.meshColor = rd.FloatVector(1.0, 0.0, 1.0, 1.0)

        # Render and readback
        output.SetMeshDisplay(cfg)
        output.Display()

        pixels = output.ReadbackOutputTexture()
        if not pixels:
            output.Shutdown()
            return ""

        # Write as PNG using PIL
        ext = ".png" if params.get("filetype", "png") == "png" else ".jpg"
        full_path = out_path + ext
        self._write_rgb_to_image(full_path, pixels, width, height)

        output.Shutdown()
        return full_path

    def _write_rgb_to_image(self, path: str, rgb_data: bytes, width: int, height: int) -> None:
        """Write raw RGB byte data as an image file."""
        from PIL import Image
        img = Image.frombytes("RGB", (width, height), rgb_data)
        if path.endswith(".jpg"):
            img.save(path, "JPEG")
        else:
            img.save(path, "PNG")

    # ------------------------------------------------------------------
    # Log — detailed text log of the result row
    # ------------------------------------------------------------------

    def _render_log(self, params: dict, event_id: int, out_path: str, row: dict) -> str:
        """Write a detailed log file for this event."""
        filetype = params.get("filetype", "txt")
        full_path = out_path + f".{filetype}"

        if filetype == "json":
            import json
            with open(full_path, "w") as f:
                json.dump(row, f, indent=2, default=str)
        else:
            lines = []
            lines.append(f"=== RenderQuery Event Log ===")
            lines.append(f"Event ID: {row.get('event_id')}")
            lines.append(f"Action Name: {row.get('name')}")
            lines.append(f"Flags: 0x{row.get('flags', 0):x}")
            lines.append(f"GPU Duration: {row.get('duration_gpu')} us")
            lines.append(f"CPU Duration: {row.get('duration_cpu')} us")
            lines.append(f"Num Indices: {row.get('num_indices')}")
            lines.append(f"Num Instances: {row.get('num_instances')}")
            lines.append(f"Draw Index: {row.get('draw_index')}")
            lines.append(f"Outputs: {row.get('outputs')}")
            lines.append(f"Depth Output: {row.get('depth_out')}")
            lines.append(f"Copy Source: {row.get('copy_source')}")
            lines.append(f"Copy Destination: {row.get('copy_destination')}")
            lines.append(f"")
            lines.append(f"=== Artifacts ===")
            for k, v in row.items():
                if k not in ("event_id", "name", "flags", "duration_gpu", "duration_cpu",
                             "num_indices", "num_instances", "draw_index", "outputs",
                             "depth_out", "copy_source", "copy_destination",
                             "custom_name", "action_id", "parent_id", "base_vertex",
                             "index_offset", "vertex_offset", "instance_offset",
                             "dispatch_dimension", "dispatch_threads_dimension",
                             "dispatch_base"):
                    lines.append(f"  {k}: {v}")
            with open(full_path, "w") as f:
                f.write("\n".join(lines))
        return full_path

    # ------------------------------------------------------------------
    # Excel — CSV/XLSX table of the result row
    # ------------------------------------------------------------------

    def _render_excel(self, params: dict, event_id: int, out_path: str, row: dict) -> str:
        """Write a CSV or XLSX file for this event."""
        filetype = params.get("filetype", "csv")
        full_path = out_path + f".{filetype}"

        if filetype == "xlsx":
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = f"Event {event_id}"
                for col, (k, v) in enumerate(row.items(), 1):
                    ws.cell(row=1, column=col, value=k)
                    ws.cell(row=2, column=col, value=str(v) if v is not None else "")
                wb.save(full_path)
                return full_path
            except ImportError:
                # Fall back to CSV if openpyxl not available
                filetype = "csv"
                full_path = out_path + ".csv"

        if filetype == "csv":
            import csv
            with open(full_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(row.keys())
                writer.writerow([str(v) if v is not None else "" for v in row.values()])
            return full_path


def _str_to_resource_id(rid_str: str) -> rd.ResourceId:
    """Parse a ResourceId from its string representation."""
    if not rid_str or rid_str == "ResourceId::Null()":
        return rd.ResourceId.Null()
    try:
        return rd.ResourceId(int(rid_str))
    except (ValueError, TypeError):
        return rd.ResourceId.Null()