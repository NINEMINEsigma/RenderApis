"""Shared Excel summary writer for renderquery examples.

``write_summary_xlsx`` renders one or more sheets from plain row dicts,
embedding image fields (screenshot paths) as scaled pictures. Falls back
to per-sheet CSV files when openpyxl is unavailable.
"""

from __future__ import annotations

import csv
import os

_IMAGE_WIDTH = 256


def write_summary_xlsx(path: str, sheets: list[dict]) -> None:
    """Write a multi-sheet summary workbook.

    Each sheet dict:
        title:        sheet name
        headers:      column header labels
        fields:       row-dict keys, one per column
        widths:       column widths in characters, one per column
        rows:         list of row dicts
        image_fields: {row-dict key: header label} — the field value is an
                      image file path embedded into the labelled column
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        _write_csv_fallback(path, sheets)
        return

    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets:
        ws = wb.create_sheet(title=spec["title"])
        _write_sheet(ws, spec)
    wb.save(path)


def _write_sheet(ws, spec: dict) -> None:
    headers = spec["headers"]
    fields = spec["fields"]
    image_fields = spec.get("image_fields", {})
    image_cols = {
        field: headers.index(label) + 1
        for field, label in image_fields.items()
        if label in headers and field in fields
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)
    for col, width in enumerate(spec.get("widths", []), 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    for i, row in enumerate(spec["rows"]):
        row_num = i + 2
        for col, field in enumerate(fields, 1):
            if field in image_cols:
                continue
            value = row.get(field, "")
            ws.cell(row=row_num, column=col, value=value if value is not None else "")
        for field, col in image_cols.items():
            img_path = row.get(field) or ""
            cell_ref = f"{ws.cell(row=row_num, column=col).column_letter}{row_num}"
            if img_path and os.path.isfile(img_path):
                _embed_image(ws, img_path, cell_ref, row_num=row_num)
            else:
                ws.cell(row=row_num, column=col, value=img_path)


def _embed_image(ws, img_path: str, cell_ref: str, target_width: int = _IMAGE_WIDTH,
                 row_num: int | None = None) -> None:
    """Embed an image scaled to target_width, growing the row height to fit."""
    from openpyxl.drawing.image import Image as XlImage
    try:
        from PIL import Image as PILImage
        with PILImage.open(img_path) as pil_img:
            orig_w, orig_h = pil_img.size
    except Exception:
        ws[cell_ref] = img_path
        return
    scale = target_width / orig_w
    img = XlImage(img_path)
    img.width = target_width
    img.height = int(orig_h * scale)
    ws.add_image(img, cell_ref)
    if row_num is not None:
        ws.row_dimensions[row_num].height = max(
            ws.row_dimensions[row_num].height or 0, int(orig_h * scale) * 0.75
        )


def _write_csv_fallback(path: str, sheets: list[dict]) -> None:
    stem, _ = os.path.splitext(path)
    for spec in sheets:
        csv_path = f"{stem}_{spec['title']}.csv"
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(spec["headers"])
            for row in spec["rows"]:
                writer.writerow([
                    "" if row.get(field) is None else str(row.get(field, ""))
                    for field in spec["fields"]
                ])
        print(f"openpyxl not available, wrote {csv_path}")