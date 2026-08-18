"""Example: Get screenshots, mesh screenshots, logs and summary Excel for ALL drawcalls.

Unlike ``top10_gpu_drawcall_screenshots`` which only exports the slowest 10%,
this example exports every drawcall in the capture, sorted by GPU duration
descending. Useful for full-frame auditing.

Usage:
    python -m renderquery.examples.all_gpu_drawcall_screenshots <capture.rdc> [--output-dir ./out/]
"""

import os
import argparse

import renderdoc as rd

from renderquery.sdk import RenderQueryClient
from renderquery.engine import artifacts


def main():
    parser = argparse.ArgumentParser(description="Full drawcall GPU-duration analysis (no filtering)")
    parser.add_argument("capture", help="Path to .rdc capture file")
    parser.add_argument("--output-dir", default="./out/", help="Output directory for artifacts")
    args = parser.parse_args()

    print(f"Opening capture: {args.capture}")
    client = RenderQueryClient(args.capture)

    print("Querying ALL drawcalls by GPU duration...")
    query = (client.query()
        .from_actions(flags=int(rd.ActionFlags.Drawcall))
        .with_gpu_counter(int(rd.GPUCounter.EventGPUDuration))
        .sort_by("duration_gpu", desc=True)
        .project(
            event_id="{event_id}",
            name="{name}",
            duration_gpu="{duration_gpu}",
            duration_cpu="{duration_cpu}",
            num_indices="{num_indices}",
            num_instances="{num_instances}",
            screenshot=artifacts.screenshot(width=512, height=512, overlay="Drawcall"),
            mesh_screenshot=artifacts.mesh_screenshot(width=512, height=512, stage="PreVS"),
            mesh=artifacts.mesh(stage="PreVS"),
            log=artifacts.log(filetype="txt"),
        )
        .to_file(args.output_dir))

    results = client.execute(query, args.output_dir)

    print(f"\nGot {len(results)} results:\n")
    for r in results:
        dur = r.get("duration_gpu")
        dur_str = f"{dur:.3f}us" if dur is not None else "N/A"
        print(f"  event {r['event_id']:>6d} | {r['name']:<40s} | GPU {dur_str}")
        print(f"    screenshot:     {r.get('screenshot', '')}")
        print(f"    mesh_screenshot: {r.get('mesh_screenshot', '')}")
        print(f"    mesh:            {r.get('mesh', '')}")
        print(f"    log:             {r.get('log', '')}")
        print()

    # Generate summary Excel with embedded images
    xlsx_path = os.path.join(args.output_dir, "summary.xlsx")
    _write_summary_xlsx(xlsx_path, results)
    print(f"Summary Excel: {xlsx_path}")

    client.shutdown()
    print("Done.")


def _write_summary_xlsx(path: str, results: list[dict]) -> None:
    """Write a summary xlsx with columns: event_id, duration, DC screenshot, mesh screenshot."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available, falling back to CSV")
        _write_summary_csv(path.replace(".xlsx", ".csv"), results)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Drawcall Analysis"

    # Header row
    headers = ["Event ID", "Name",
               "GPU Duration (GPU执行耗时, us)", "CPU Duration (API提交耗时, us)",
               "Num Indices", "DC Screenshot", "Mesh Screenshot"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30

    for i, r in enumerate(results):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=r.get("event_id", ""))
        ws.cell(row=row_num, column=2, value=r.get("name", ""))
        dur = r.get("duration_gpu")
        ws.cell(row=row_num, column=3, value=float(dur) if dur is not None else "")
        ws.cell(row=row_num, column=4, value=r.get("duration_cpu", ""))
        ws.cell(row=row_num, column=5, value=r.get("num_indices", ""))

        # Embed DC screenshot (preserve aspect ratio)
        screenshot_path = r.get("screenshot", "")
        if screenshot_path and os.path.isfile(screenshot_path):
            try:
                from PIL import Image as PILImage
                pil_img = PILImage.open(screenshot_path)
                orig_w, orig_h = pil_img.size
                pil_img.close()
                # Fit to 256px wide, scale height proportionally
                scale = 256 / orig_w
                img = XlImage(screenshot_path)
                img.width = 256
                img.height = int(orig_h * scale)
                ws.add_image(img, f"F{row_num}")
                ws.row_dimensions[row_num].height = max(ws.row_dimensions[row_num].height or 0, int(orig_h * scale) * 0.75)
            except Exception:
                ws.cell(row=row_num, column=6, value=screenshot_path)

        # Embed mesh screenshot (preserve aspect ratio)
        mesh_path = r.get("mesh_screenshot", "")
        if mesh_path and os.path.isfile(mesh_path):
            try:
                from PIL import Image as PILImage
                pil_img2 = PILImage.open(mesh_path)
                orig_w2, orig_h2 = pil_img2.size
                pil_img2.close()
                scale2 = 256 / orig_w2
                img2 = XlImage(mesh_path)
                img2.width = 256
                img2.height = int(orig_h2 * scale2)
                ws.add_image(img2, f"G{row_num}")
            except Exception:
                ws.cell(row=row_num, column=7, value=mesh_path)

    wb.save(path)


def _write_summary_csv(path: str, results: list[dict]) -> None:
    """Fallback CSV without images."""
    import csv
    if not results:
        return
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=results[0].keys())
        writer.writeheader()
        for r in results:
            writer.writerow({k: str(v) if v is not None else "" for k, v in r.items()})


if __name__ == "__main__":
    main()